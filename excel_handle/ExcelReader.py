from dataclasses import fields
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from .schema import SourceFieldMeta, GroupLevel
from typing import Any, List, Tuple, Union

class MyExcelReader():
    def __init__(self, filename:str, aFieldStr:List[SourceFieldMeta]): #'C://Users//BhargavaTanguturi//Downloads//Journal.xlsx'
        self._book:Workbook = load_workbook(filename)
        self._aFieldStr = aFieldStr
    
    def _get_sheet(self, sSheetName:str) -> Worksheet:
        if sSheetName:
            return self._book.get_sheet_by_name(sSheetName)
        else:
            return self._book.active

    def _separate_fields(self, aFieldStr:List[SourceFieldMeta]) -> Tuple[List[SourceFieldMeta]]:
        aGroupedFieldsAsForm:List[SourceFieldMeta] = []
        aGroupedFields:List[SourceFieldMeta] = []
        aFlatFields:List[SourceFieldMeta] = []
        for oField in aFieldStr:
            if oField.IsGrouped == True and oField.GroupedAndHeaderAsForm == True:
                aGroupedFieldsAsForm.append(oField)
            elif oField.IsGrouped == True:
                aGroupedFields.append(oField)
            else:
                aFlatFields.append(oField)
        
        return (aGroupedFieldsAsForm, aGroupedFields, aFlatFields)
    


    def _determine_columns(self, oSheet:Worksheet, aFieldStr:List[SourceFieldMeta]):
        aCollectedFields:List[SourceFieldMeta] = []
        bColumnsRepeated = False
        nFormColumnStartRowIndex:int = 0
        nTableColumnStartRowIndex:int = 0
        # for row in oSheet.rows:
        bIsTableHeaderRow = False
        for row in oSheet.iter_rows(max_row=20):
            oField:SourceFieldMeta = None
            if bIsTableHeaderRow:
                break
            for cell in row:
                if isinstance(cell, MergedCell): # this cell Merged with previous cell 
                    if oField:
                        oField.MergedColumns = oField.MergedColumns + 1
                    continue
                
                cell:Cell = cell
                if oField and oField.GroupedAndHeaderAsForm and oField.FieldValueIndex < 1: # this would be value cell of form based field
                    oField.FieldValueIndex = cell.col_idx
                    continue
                else:
                    oField:SourceFieldMeta = None
                    # oField = _find_record(aCollectedFields, "SourceFieldName", cell.value)
                    for fld in aCollectedFields:
                        if fld.SourceFieldName == cell.value and fld.GroupedAndHeaderAsForm != bIsTableHeaderRow:
                            oField = fld
                            break
                    if oField:
                        bColumnsRepeated = True
                        continue
                    # oField = _find_record(aFieldStr, "SourceFieldName", cell.value)
                    for fld in aFieldStr:
                        if bIsTableHeaderRow:
                            if fld.SourceFieldName == cell.value and fld.GroupedAndHeaderAsForm != bIsTableHeaderRow:
                                oField = fld
                                break
                        elif fld.SourceFieldName == cell.value:
                            oField = fld
                            break
                    if oField:
                        oField.FieldNameIndex = cell.col_idx
                        if oField.GroupedAndHeaderAsForm:
                            nFormColumnStartRowIndex = cell.row if nFormColumnStartRowIndex == 0 else nFormColumnStartRowIndex
                        else:
                            bIsTableHeaderRow = True
                            oField.FieldValueIndex = cell.col_idx
                            nTableColumnStartRowIndex = cell.row if nTableColumnStartRowIndex == 0 else nTableColumnStartRowIndex

                        aCollectedFields.append(oField)
                    if len(aCollectedFields) == len(aFieldStr):
                        bColumnsRepeated = True
                        break
            if bColumnsRepeated:
                break
        
        return (aCollectedFields, nFormColumnStartRowIndex, nTableColumnStartRowIndex)


    def get_sheet_data(self, sSheetName:str=None, aFieldStr:List[SourceFieldMeta]=[]):
        sheet = self._get_sheet(sSheetName=sSheetName)
        # (aCollectedFields, nFormColumnStartRowIndex, nTableColumnStartRowIndex) = self._determine_columns(sheet, aFieldStr)
        return self.collect_records(sheet, *self._determine_columns(sheet, aFieldStr))
        # (aGroupedFieldsAsForm, aGroupedFields, aFlatFields) = self._separate_fields(aCollectedFields)

    def prepare_first_level_form_fields(self, aFormFields:List[SourceFieldMeta])->GroupLevel:
        oFormLevel=GroupLevel(Level=0, Fields=[])
        if aFormFields:
            oFormLevel = GroupLevel(Level=0, Fields= [], MustBeField=None, FirstField=None, LastField=None)
            for fld in aFormFields:
                oFormLevel.Fields.append(fld)
                if fld.IsFormStart:
                    oFormLevel.FirstField = fld
                if fld.IsFormEnd:
                    oFormLevel.LastField = fld
        
        return oFormLevel

    def prepare_grouped_fields_by_level(self, aGroupedFields:List[SourceFieldMeta]) -> List[GroupLevel]:
        aFieldsByLevel = []
        for oGroupField in aGroupedFields:
            oLevel:GroupLevel = _find_record(aFieldsByLevel, "Level", oGroupField.GroupedLevel)
            if not oLevel:
                oLevel = GroupLevel(Level= oGroupField.GroupedLevel, Fields= [], MustBeField=oGroupField)
                aFieldsByLevel.append(oLevel)
            oLevel.Fields.append(oGroupField)
        
        return aFieldsByLevel

    def prepare_last_level_fields(self, aFlatFields:List[SourceFieldMeta], nGroupedLevelsCount:int = 1) -> GroupLevel:
        oLastLevel = GroupLevel(Level= nGroupedLevelsCount + 1, Fields= [])
        if aFlatFields:
            for oField in aFlatFields:
                oLastLevel.Fields.append(oField)
                if oField.ShouldBe:
                    oLastLevel.MustBeField = oField
        return oLastLevel

    def collect_form_based_values(self, row:Tuple[Union[Cell, MergedCell]], oFormLevel:GroupLevel) -> dict:
        form_record = {}
        for oFormField in oFormLevel.Fields:
            sFieldName = row[oFormField.FieldNameIndex - 1].value
            if sFieldName == oFormField.SourceFieldName:
                form_record[oFormField.AliasName] = row[oFormField.FieldValueIndex - 1].value
        return form_record

    def collect_row_group_based_values(self, row:Tuple[Union[Cell,MergedCell]], oLevel:GroupLevel) -> Tuple[dict, bool]:
        bInValidData = False
        group_record = {}
        for fld in oLevel.Fields:
            sCellValue = row[fld.FieldValueIndex - 1].value
            if not sCellValue and fld.Required:
                bInValidData = True
            else:
                group_record[fld.AliasName] = sCellValue
        
        return (group_record, bInValidData)
    
        
        



    def collect_records(self, sheet:Worksheet, aCollectedFields:List[SourceFieldMeta], nFormColumnStartRowIndex:int, nTableColumnStartRowIndex:int):
        (aGroupedFieldsAsForm, aGroupedFields, aFlatFields) = self._separate_fields(aCollectedFields)
        aData= []
        oField:SourceFieldMeta
        oLevel:GroupLevel
        oFormLevel = self.prepare_first_level_form_fields(aGroupedFieldsAsForm)
        bIncludeFormGroup = True if oFormLevel.FirstField is not None else False 
        
        aFieldsByGroupLevel = self.prepare_grouped_fields_by_level(aGroupedFields)
        oLastLevel= self.prepare_last_level_fields(aFlatFields, nGroupedLevelsCount=len(aFieldsByGroupLevel))

        aFieldsByGroupLevel.sort(key=lambda x: x.Level)
        aRecordByLevels = {}

        bInvalidGroupLevelData = True
        bFormStarted = False
        bFormEnd = False
        bTableStarted = True
        bFindNextTableHeader = True
        form_record = {}
        row:List[Union[Cell,MergedCell]]
        for row in sheet.rows:
            if bIncludeFormGroup:
                if oFormLevel.FirstField and row[oFormLevel.FirstField.FieldNameIndex -1].value == oFormLevel.FirstField.SourceFieldName:
                   bFormStarted = True
                   bFormEnd = False
                   bFindNextTableHeader = False
                   bTableStarted = False
                   form_record = {}
                
                if bFormStarted:
                    form_record = {**form_record, **self.collect_form_based_values(row=row, oFormLevel=oFormLevel)}
                
                if oFormLevel.LastField and row[oFormLevel.LastField.FieldNameIndex -1].value == oFormLevel.LastField.SourceFieldName:
                    bFormEnd = True
                    bFormStarted = True
                    bFindNextTableHeader = True
                    continue
            else: 
                bFormEnd = True
            
            if not bFormEnd:
                continue

            if bFindNextTableHeader:
                if bIncludeFormGroup:
                    if oLastLevel.MustBeField and row[oLastLevel.MustBeField.FieldNameIndex - 1].value == oLastLevel.MustBeField.SourceFieldName:
                        bFindNextTableHeader = False
                        bTableStarted = True
                        continue
                elif row[0].row <= nTableColumnStartRowIndex: # data start index
                    continue
                else:
                    bTableStarted = True

            if bTableStarted:      
                group_record = {}
                record = {}
                bValidRecord = True
                if aFieldsByGroupLevel:
                    for oLevel in aFieldsByGroupLevel: # collect grouped fields data order by level
                        # the field must have value at group level otherwise it's group level data
                        if oLevel.MustBeField and row[oLevel.MustBeField.FieldValueIndex - 1].value:  
                            (group_record, bInvalidGroupLevelData) =self.collect_row_group_based_values(row=row, oLevel=oLevel)
                        
                            aRecordByLevels[oLevel.Level]= group_record
                else:
                    bInvalidGroupLevelData = False # no grouped fields

                if bInvalidGroupLevelData: # skip row due to missing or invalid data in grouped fields
                    continue # skip up to next valid group record

                if oLastLevel:
                    if oLastLevel.MustBeField and row[oLastLevel.MustBeField.FieldValueIndex - 1].value:
                        for oField in oLastLevel.Fields:
                            record[oField.AliasName] = row[oField.FieldValueIndex - 1].value
                            if oField.Required and record[oField.AliasName] is None:
                                bValidRecord = False

                        if bValidRecord:
                            for sLevelKey in aRecordByLevels:
                                record = {**aRecordByLevels[sLevelKey], **record}
                            record = {**form_record, **record}
                            aData.append(record) 

        return aData

def _find_record(aArray:List[Any], sFieldName:str, sValue:Any):
    bExist = False
    for record in aArray:
        if isinstance(record, dict):
            sFieldValue = record.get(sFieldName)
        else:
            sFieldValue = getattr(record, sFieldName)
        if sFieldValue == sValue:
            bExist = True
            break
    
    return record if bExist == True else None